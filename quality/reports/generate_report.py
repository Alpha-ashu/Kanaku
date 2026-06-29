import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registration QA Report"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    body_font = Font(name=font_family, size=10)
    pass_font = Font(name=font_family, size=10, bold=True, color="006100")
    fail_font = Font(name=font_family, size=10, bold=True, color="9C0006")
    
    title_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    # Title Row
    ws.merge_cells("A1:H1")
    ws["A1"] = "KANAKU - USER REGISTRATION FEATURE QA VALIDATION REPORT"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Header Row
    headers = [
        "Test Case ID", "Test Category", "Scenario / Description", 
        "Prerequisites", "Test Data", "Expected Result", 
        "Actual Result", "Status"
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[2].height = 28
    
    # Data Rows
    scenarios = [
        # UI
        ("TC-REG-UI-01", "UI Testing", "Happy path: complete registration flow on UI and redirects to onboarding", "No active session, Dev server running", "firstName: Dynamic, lastName: Dynamic, email: Dynamic, mobile: Dynamic, password: Password123!", "Account created successfully, user is logged in, and onboarding flow is shown.", "User successfully registered, token saved in local storage, redirected to onboarding flow.", "PASS"),
        ("TC-REG-UI-02", "UI Testing", "UI Validation: blocks registration with invalid email format", "Dev server running", "firstName: Dynamic, lastName: Dynamic, email: not-an-email, mobile: Dynamic, password: Password123!", "Submit button is disabled, inline error 'Invalid email address' is displayed.", "Submit button was disabled, inline error displayed under Email Address field.", "PASS"),
        ("TC-REG-UI-03", "UI Testing", "UI Validation: blocks registration with mismatched passwords", "Dev server running", "email: Dynamic, password: Password123!, confirmPassword: DifferentPassword123!", "Submit button is disabled, mismatched password warning is shown.", "Submit button remained disabled due to password validation mismatch.", "PASS"),
        ("TC-REG-UI-04", "UI Testing", "UI Validation: blocks registration when terms are unchecked", "Dev server running", "All valid inputs, terms checkbox: unchecked", "Submit button is disabled.", "Submit button remained disabled until terms checkbox was checked.", "PASS"),
        ("TC-REG-UI-05", "UI Testing", "UI Validation: duplicate email is blocked with generic message, no success screen", "Dev server running, email already registered", "email: existing_user@kanaku.test", "Inline error 'This email can't be used for a new account' is displayed on blur; submit disabled.", "Email blur triggered unique email check, showing generic inline rejection, submit button disabled.", "PASS"),
        # API
        ("TC-REG-API-01", "API Testing", "registers a new user successfully and creates DB records (201)", "Prisma client initialized", "name: Dynamic, email: Dynamic, password: Password123!, mobile: Dynamic", "Returns HTTP 201 Created with JSON response containing user object, access token, and refresh token cookie.", "HTTP 201 returned with correct JSON structure, user record written to DB.", "PASS"),
        ("TC-REG-API-02", "API Testing", "rejects duplicate email with 409 Conflict", "Email already exists in database", "email: existing_user@kanaku.test", "Returns HTTP 409 Conflict with EMAIL_EXISTS error code.", "HTTP 409 Conflict returned with EMAIL_EXISTS code.", "PASS"),
        ("TC-REG-API-03", "API Testing", "rejects duplicate mobile/phone with 409 Conflict", "Phone already exists in database", "mobile: 9999999999", "Returns HTTP 409 Conflict with PHONE_EXISTS error code.", "HTTP 409 Conflict returned with PHONE_EXISTS code.", "PASS"),
        ("TC-REG-API-04", "API Testing", "rejects validation scenario: Empty Email", "Dev server running", "email: ''", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400 Bad Request returned with validation details.", "PASS"),
        ("TC-REG-API-05", "API Testing", "rejects validation scenario: Invalid Email Format", "Dev server running", "email: invalid-format", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400 Bad Request returned with validation details.", "PASS"),
        ("TC-REG-API-06", "API Testing", "rejects validation scenario: Missing Name", "Dev server running", "name: absent/empty", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400/Bad Request returned.", "PASS"),
        ("TC-REG-API-07", "API Testing", "rejects validation scenario: Short Password", "Dev server running", "password: short", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400/Bad Request returned.", "PASS"),
        ("TC-REG-API-08", "API Testing", "rejects validation scenario: Weak Password - No special chars", "Dev server running", "password: WeakPassword123", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400/Bad Request returned.", "PASS"),
        ("TC-REG-API-09", "API Testing", "rejects validation scenario: Weak Password - No uppercase", "Dev server running", "password: weakpassword123!", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400/Bad Request returned.", "PASS"),
        ("TC-REG-API-10", "API Testing", "rejects validation scenario: Weak Password - No digits", "Dev server running", "password: WeakPassword!", "Returns HTTP 400 Bad Request with validation error code.", "HTTP 400/Bad Request returned.", "PASS"),
        ("TC-REG-API-11", "API Testing", "sanitizes and processes Name with SQL Injection payload safely", "Dev server running", "name: ' OR '1'='1", "Returns HTTP 201 Created. DB stores raw payload safely without executing the injection.", "HTTP 201 returned. Name stored as literal string in database, verified secure.", "PASS"),
        ("TC-REG-API-12", "API Testing", "processes Name with XSS payload safely by stripping scripts", "Dev server running", "name: QA User <script>alert(\"xss\")</script>", "Returns HTTP 201 Created. Script tags are stripped from the name attribute before writing to DB.", "HTTP 201 returned. Script tag stripped, name saved as 'QA User' in database.", "PASS"),
        # DB
        ("TC-REG-DB-01", "Database Validation", "verifies registration inputs correctly map to DB tables and reflect in profile endpoints", "Unique user credentials", "All registration parameters", "Rows added to User, profiles, UserSettings, and default Account tables with correct matching values.", "All database tables verify exact column mapping and correct relation constraints.", "PASS"),
        # Forgot Password
        ("TC-REG-FPS-01", "Forgot Password", "happy path: request OTP, verify and reset password successfully", "Registered user exists", "email: Dynamic, newPassword: NewCoolPassword123!", "User can request OTP, verify it, reset password, and log in with new password. Old password is deleted.", "Reset OTP requested, verified, password updated to new hashed value, login successful.", "PASS")
    ]
    
    for row_idx, data in enumerate(scenarios, 3):
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
            
            # Status styling
            if col_idx == 8:
                if val == "PASS":
                    cell.font = pass_font
                    cell.fill = pass_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif val == "FAIL":
                    cell.font = fail_font
                    cell.fill = fail_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 36
        
    # Auto-adjust column widths with safety padding
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val_str = str(cell.value or '')
            # If cell has multiple lines, take the longest line
            lines = val_str.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
        
    wb.save("Kanaku_Registration_QA_Report.xlsx")
    print("Report created successfully: Kanaku_Registration_QA_Report.xlsx")

if __name__ == "__main__":
    create_report()
